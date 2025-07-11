
import openpyxl

def read_excel_sheets(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames
    
    all_sheet_data = {}
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        all_sheet_data[sheet_name] = data
    
    return all_sheet_data

file_path = '/home/ubuntu/upload/DASHBOARDHARGI.xlsx'
sheet_data = read_excel_sheets(file_path)

# Print more rows for 'REALISASI HAR GI' and 'drop list' sheets
for sheet_name, data in sheet_data.items():
    print(f'Sheet: {sheet_name}')
    if sheet_name == 'REALISASI HAR GI' or sheet_name == 'drop list':
        for row in data[:10]:  # Print first 10 rows
            print(row)
    else:
        for row in data[:5]:  # Print first 5 rows for other sheets
            print(row)
    print('\n')


